from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

excel_template_path = r'C:\Users\TI DEV\simplex_abdias\GERAR OP SERRA.xlsx'  # Substitua pelo caminho do arquivo de template

# Função para preencher a planilha Excel com os resultados da ordem
def preencher_excel_ordem(excel_template_path, resultados_iteracao, ordem_global):

    # Carrega o template Excel
    wb = load_workbook(excel_template_path)
    ws = wb.active  # Supondo que o template tem apenas uma planilha ativa

    start_row = 9  # Linha inicial para preenchimento
    codigo_col = 'B'  # Coluna para os códigos
    quantidade_col = 'Z'  # Coluna para as quantidades
    tamanho_col = 'AD'  # Coluna para os comprimentos

    # Preenche as células com os dados
    for idx, item in enumerate(resultados_iteracao):
        row = start_row + idx
        ws[f'{codigo_col}{row}'] = item['codigo']
        ws[f'{quantidade_col}{row}'] = item['quantidade_alocada']
        ws[f'{tamanho_col}{row}'] = item['tamanho_alocado_item']

    # Salva a planilha com um nome específico para a ordem
    output_excel_path = f'OP{ordem_global}.xlsx'
    wb.save(output_excel_path)
    print(f"Ordem OP{ordem_global} salva em {output_excel_path}")