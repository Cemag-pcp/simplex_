import csv
import pulp
import pandas as pd
import time
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pickle  # For storing results

def tratar(input_csv):
    df = pd.read_csv(input_csv, sep=',', encoding='UTF-8')

    # df = df.drop(columns={'Unnamed: 0'})
    # df['CODIGO'] = df['DESCRIÇÃO'].str.split(' - ')[0][0]
    df.rename(columns={'Código': 'codigo', 'Matéria Prima': 'mp', 'Qntd': 'qtd_maxima', 'Conjunto': 'conjuntos', 'Comprimento': 'tamanhos'}, inplace=True)
    df.dropna(inplace=True)

    # Converter 'qtd_maxima' para float
    df['qtd_maxima'] = df['qtd_maxima'].astype(str).str.replace('.', '').str.replace(',', '.').astype(float)

    # Converter 'tamanhos' para float
    df['tamanhos'] = df['tamanhos'].astype(str).str.replace('.', '').str.replace(',', '.').astype(float)

    df_agrouped = df.groupby(['conjuntos','mp','codigo']).agg({
        'qtd_maxima': 'sum',
        'tamanhos': 'mean'
    }).reset_index().sort_values('mp')

    return df_agrouped

def carregar_dados_csv(input_csv):
    df = tratar(input_csv)

    df['codigo'] = df['codigo'].astype(str)
    df['qtd_maxima'] = df['qtd_maxima'].astype(float)
    df['tamanhos'] = df['tamanhos'].astype(float)
    df['conjuntos'] = df['conjuntos'].astype(str)
    df['mp'] = df['mp'].astype(str)

    return df

# Função para resolver o problema de otimização e armazenar os resultados
def otimizar_e_armazenar_resultados(df, resultados_pickle_path, resultados_csv_path):
    ordem_global = 192  # Variável para manter o controle das ordens globais
    resultados_totais = []  # Lista para armazenar todos os resultados

    # Agrupar o DataFrame por 'conjunto' e 'mp'
    grupos = df.groupby(['mp'])

    # Iterar sobre cada grupo
    for mp, grupo in grupos:
        # print(f"\nProcessando MP: {mp}")
        tamanho_total = (grupo['qtd_maxima'] * grupo['tamanhos']).sum()

        # Exibir o resultado para cada grupo
        print(f"\nProcessando MP: {mp}")
        print(f"Tamanho total do grupo: {tamanho_total}")

        codigos_pecas = grupo['codigo'].tolist()
        qtd_maxima = grupo['qtd_maxima'].tolist()
        tamanhos = grupo['tamanhos'].tolist()
        conjuntos = grupo['conjuntos'].tolist()
        ordem = 1  # Reinicia a contagem de ordens para cada grupo

        # Loop até que todas as quantidades sejam zero para o grupo atual
        while any(qtd > 0 for qtd in qtd_maxima):
            resultados_iteracao = []
            total_tamanho_alocado_ordem = 0  # Inicializa o total alocado na ordem atual

            # Criação do problema de otimização (Minimização da folga)
            prob = pulp.LpProblem("Minimizar Folga", pulp.LpMinimize)

            # Definindo as variáveis de decisão com limites baseados na quantidade máxima atual
            variaveis = []
            for i in range(len(codigos_pecas)):
                nome_var = f"{codigos_pecas[i]}_{mp}_{ordem}"
                variavel = pulp.LpVariable(nome_var, lowBound=0, upBound=qtd_maxima[i], cat='Integer')
                variaveis.append(variavel)

            # Variável de folga (diferença entre 6000 e o total alocado)
            folga = pulp.LpVariable('Folga', lowBound=0, cat='Continuous')

            # Definindo a função objetivo (minimizar a folga)
            prob += folga, "Folga Total"

            # Restrição: total alocado + folga = 6000
            prob += pulp.lpSum([variaveis[i] * tamanhos[i] for i in range(len(codigos_pecas))]) + folga == tamanho_total, "Restrição de tamanho total alocado"

            # Resolvendo o problema
            prob.solve()

            # Verificar se uma solução viável foi encontrada
            if prob.status != pulp.LpStatusOptimal:
                print(f"Não foi possível encontrar uma solução ótima na Ordem {ordem} para MP {mp}. Encerrando este grupo.")
                break

            # Armazenando os valores das variáveis de decisão (quantidade alocada de cada item) se for maior que 0
            for i, variable in enumerate(variaveis):
                if variable.varValue and variable.varValue > 0:
                    quantidade_alocada = int(variable.varValue)
                    tamanho_alocado_item = quantidade_alocada * tamanhos[i]
                    total_tamanho_alocado_ordem += tamanho_alocado_item
                    resultados_iteracao.append({
                        'mp': mp,
                        'conjunto': conjuntos[i],
                        'codigo': codigos_pecas[i],
                        'quantidade_alocada': quantidade_alocada,
                        'tamanho_alocado_item': tamanho_alocado_item,
                        'ordem_global': ordem_global,
                        'ordem_grupo': ordem,
                        'perda_materia_prima': None  # Será calculado depois
                    })
                    # Reduzindo a quantidade máxima da peça correspondente
                    qtd_maxima[i] -= quantidade_alocada

            # Evitando valores negativos
            qtd_maxima = [max(0, qtd) for qtd in qtd_maxima]

            # Calcula a perda de matéria-prima para esta ordem
            perda_materia_prima = tamanho_total - total_tamanho_alocado_ordem

            # Atualiza a perda de matéria-prima em cada item da iteração atual
            for item in resultados_iteracao:
                item['perda_materia_prima'] = perda_materia_prima

            # Armazenando os resultados da iteração atual nos resultados totais
            resultados_totais.extend(resultados_iteracao)

            print(f"Ordem {ordem} processada para MP {mp}. Tamanho total alocado: {total_tamanho_alocado_ordem}.")
            ordem += 1
            ordem_global += 1  # Incrementa a ordem global

    # Armazenar os resultados em um arquivo pickle para uso posterior
    with open(resultados_pickle_path, 'wb') as f:
        pickle.dump(resultados_totais, f)

    # Converter a lista de resultados em um DataFrame
    df_resultados = pd.DataFrame(resultados_totais)

    # Reorganizar as colunas e salvar em CSV
    colunas = ['ordem_global', 'ordem_grupo', 'conjunto', 'mp', 'codigo', 'quantidade_alocada', 'tamanho_alocado_item', 'perda_materia_prima']
    df_resultados = df_resultados[colunas]
    df_resultados.to_csv(resultados_csv_path, index=False)

    print(f"\nProcessamento concluído. Resultados armazenados em {resultados_pickle_path} e {resultados_csv_path}.")

# Função para gerar os arquivos Excel com base nos resultados armazenados
def gerar_arquivos_excel(resultados_pickle_path, excel_template_path):
    # Carregar os resultados armazenados
    with open(resultados_pickle_path, 'rb') as f:
        resultados_totais = pickle.load(f)

    # Agrupar os resultados por ordem
    df_resultados = pd.DataFrame(resultados_totais)
    grupos_ordem = df_resultados.groupby('ordem_global')

    for ordem, grupo in grupos_ordem:
        resultados_iteracao = grupo.to_dict('records')
        # Preencher o Excel com os resultados da ordem atual
        preencher_excel_ordem(excel_template_path, resultados_iteracao, ordem)

# Função para preencher a planilha Excel com os resultados da ordem
def preencher_excel_ordem(excel_template_path, resultados_iteracao, ordem_global):
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    # Carrega o template Excel
    wb = load_workbook(excel_template_path)
    ws = wb.active  # Supondo que o template tem apenas uma planilha ativa

    start_row = 10  # Linha inicial para preenchimento
    codigo_col = 'A'  # Coluna para os códigos
    quantidade_col = 'C'  # Coluna para as quantidades
    tamanho_col = 'D'  # Coluna para os comprimentos
    conjunto_col = 'E'
    
    # Extrai 'mp' e 'conjunto' do primeiro item
    mp = resultados_iteracao[0]['mp']
    # conjunto = resultados_iteracao[0]['conjunto']

    # Define as células especificadas
    ws['B4'] = f'OP{int(ordem_global)}'
    ws['B5'] = mp[0]
    # ws['E4'] = conjunto
    total_tamanho_alocado_ordem = sum(item['tamanho_alocado_item'] for item in resultados_iteracao)
    num_varas = -(-total_tamanho_alocado_ordem // 6000)  # Equivalente a math.ceil()
    perda_materia_prima = (num_varas * 6000) - total_tamanho_alocado_ordem

    ws['B6'] = f"{total_tamanho_alocado_ordem}mm ou {total_tamanho_alocado_ordem / 6000:.2f} varas"
    ws['E6'] = perda_materia_prima  # Exibe a perda de material em 'E6'

    # Preenche as células com os dados
    for idx, item in enumerate(resultados_iteracao):
        row = start_row + idx
        ws[f'{codigo_col}{row}'] = "0"+item['codigo'] if len(item['codigo'])==5 else item['codigo']
        ws[f'{quantidade_col}{row}'] = item['quantidade_alocada']
        ws[f'{tamanho_col}{row}'] = item['tamanho_alocado_item'] / item['quantidade_alocada']
        ws[f'{conjunto_col}{row}'] = item['conjunto']

    # Salva a planilha com um nome específico para a ordem
    output_excel_path = f'OP{int(ordem_global)}.xlsx'
    wb.save(r"C:\Users\pcp2\Downloads\teste_csv" + output_excel_path )
    print(f"Ordem OP{int(ordem_global)} salva em {output_excel_path}")

# Caminho para o CSV de entrada fornecido pelo usuário
input_csv = 'input_pecas.csv'  # Substitua pelo caminho do seu arquivo CSV de entrada

# Caminho para o template Excel fornecido pelo usuário
excel_template_path = r'C:\Users\pcp2\simplex\simplex_\GERAR OP SERRA.xlsx'  # Substitua pelo caminho do seu template Excel

# Caminho para armazenar os resultados da otimização
resultados_pickle_path = 'resultados_otimizacao.pkl'
resultados_csv_path = 'resultado_otimizacao.csv'  # Novo arquivo CSV com as informações gerais

# Carregar os dados do CSV fornecido pelo usuário
df = carregar_dados_csv(input_csv)

# Executar a otimização e armazenar os resultados (sem gerar os arquivos Excel)
otimizar_e_armazenar_resultados(df, resultados_pickle_path, resultados_csv_path)

# Em algum momento posterior, chamar a função para gerar os arquivos Excel
gerar_arquivos_excel(resultados_pickle_path, excel_template_path)
