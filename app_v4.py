import pulp
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pickle  # For storing results
import streamlit as st
from io import BytesIO
import zipfile

def tratar(df):
    # df = pd.read_csv(input_csv, sep=',', encoding='UTF-8')

    # df = df.drop(columns={'Unnamed: 0'})
    # df['CODIGO'] = df['DESCRIÇÃO'].str.split(' - ')[0][0]
    df.rename(columns={'Código': 'codigo', 'MP': 'mp', 'Qntd': 'qtd_maxima', 'Conjunto': 'conjuntos', 'Comprimento': 'tamanhos'}, inplace=True)
    df.dropna(inplace=True)

    # Converter 'qtd_maxima' para float
    df['qtd_maxima'] = df['qtd_maxima'].astype(str).str.replace('.', '').str.replace(',', '.').astype(float)

    # Converter 'tamanhos' para float
    df['tamanhos'] = df['tamanhos'].astype(str).str.replace('.', '').str.replace(',', '.').astype(float)

    df_agrouped = df.groupby(['conjuntos','mp','codigo']).agg({
        'qtd_maxima': 'sum',
        'tamanhos': 'mean'
    }).reset_index()

    return df_agrouped

def carregar_dados_csv(df):
    df = tratar(df)

    df['codigo'] = df['codigo'].astype(str)
    df['qtd_maxima'] = df['qtd_maxima'].astype(float)
    df['tamanhos'] = df['tamanhos'].astype(float)
    df['conjuntos'] = df['conjuntos'].astype(str)
    df['mp'] = df['mp'].astype(str)
    return df

# Função para resolver o problema de otimização e armazenar os resultados
def otimizar_e_armazenar_resultados(df, resultados_pickle_path, resultados_csv_path):
    ordem_global = 192  # Variável para manter o controle das ordens globais
    resultados_totais = []  # Lista para armazenar todos os resultados

    # Agrupar o DataFrame por 'conjunto' e 'mp'
    grupos = df.groupby(['mp'])

    # Iterar sobre cada grupo
    for mp, grupo in grupos:
        print(f"\nProcessando MP: {mp}")
        codigos_pecas = grupo['codigo'].tolist()
        qtd_maxima = grupo['qtd_maxima'].tolist()
        tamanhos = grupo['tamanhos'].tolist()
        conjuntos = grupo['conjuntos'].tolist()
        ordem = 1  # Reinicia a contagem de ordens para cada grupo

        # Loop até que todas as quantidades sejam zero para o grupo atual
        while any(qtd > 0 for qtd in qtd_maxima):
            resultados_iteracao = []
            total_tamanho_alocado_ordem = 0  # Inicializa o total alocado na ordem atual

            # Criação do problema de otimização (Minimização da folga)
            prob = pulp.LpProblem("Minimizar Folga", pulp.LpMinimize)

            # Definindo as variáveis de decisão com limites baseados na quantidade máxima atual
            variaveis = []
            for i in range(len(codigos_pecas)):
                nome_var = f"{codigos_pecas[i]}_{mp}_{ordem}"
                variavel = pulp.LpVariable(nome_var, lowBound=0, upBound=qtd_maxima[i], cat='Integer')
                variaveis.append(variavel)

            # Variável de folga (diferença entre 6000 e o total alocado)
            folga = pulp.LpVariable('Folga', lowBound=0, cat='Continuous')

            # Definindo a função objetivo (minimizar a folga)
            prob += folga, "Folga Total"

            # Restrição: total alocado + folga = 6000
            prob += pulp.lpSum([variaveis[i] * tamanhos[i] for i in range(len(codigos_pecas))]) + folga == 6000, "Restrição de tamanho total alocado"

            # Resolvendo o problema
            prob.solve()

            # Verificar se uma solução viável foi encontrada
            if prob.status != pulp.LpStatusOptimal:
                print(f"Não foi possível encontrar uma solução ótima na Ordem {ordem} para MP {mp}. Encerrando este grupo.")
                break

            # Armazenando os valores das variáveis de decisão (quantidade alocada de cada item) se for maior que 0
            for i, variable in enumerate(variaveis):
                if variable.varValue and variable.varValue > 0:
                    quantidade_alocada = int(variable.varValue)
                    tamanho_alocado_item = quantidade_alocada * tamanhos[i]
                    total_tamanho_alocado_ordem += tamanho_alocado_item
                    resultados_iteracao.append({
                        'mp': mp,
                        'conjunto': conjuntos[i],
                        'codigo': codigos_pecas[i],
                        'quantidade_alocada': quantidade_alocada,
                        'tamanho_alocado_item': tamanho_alocado_item,
                        'ordem_global': ordem_global,
                        'ordem_grupo': ordem,
                        'perda_materia_prima': None  # Será calculado depois
                    })
                    # Reduzindo a quantidade máxima da peça correspondente
                    qtd_maxima[i] -= quantidade_alocada

            # Evitando valores negativos
            qtd_maxima = [max(0, qtd) for qtd in qtd_maxima]

            # Calcula a perda de matéria-prima para esta ordem
            perda_materia_prima = 6000 - total_tamanho_alocado_ordem

            # Atualiza a perda de matéria-prima em cada item da iteração atual
            for item in resultados_iteracao:
                item['perda_materia_prima'] = perda_materia_prima

            # Armazenando os resultados da iteração atual nos resultados totais
            resultados_totais.extend(resultados_iteracao)

            print(f"Ordem {ordem} processada para MP {mp}. Tamanho total alocado: {total_tamanho_alocado_ordem}.")
            ordem += 1
            ordem_global += 1  # Incrementa a ordem global

    # Armazenar os resultados em um arquivo pickle para uso posterior
    with open(resultados_pickle_path, 'wb') as f:
        pickle.dump(resultados_totais, f)

    # Converter a lista de resultados em um DataFrame
    df_resultados = pd.DataFrame(resultados_totais)

    # Reorganizar as colunas e salvar em CSV
    colunas = ['ordem_global', 'ordem_grupo', 'conjunto', 'mp', 'codigo', 'quantidade_alocada', 'tamanho_alocado_item', 'perda_materia_prima']
    df_resultados = df_resultados[colunas]
    df_resultados.to_csv(resultados_csv_path, index=False)

    print(f"\nProcessamento concluído. Resultados armazenados em {resultados_pickle_path} e {resultados_csv_path}.")

def consolidar_codigos_quantidades(grupo, qtd_vara):
    codigos_quantidades = {}
    for codigo, qtd, comprimento, conjunto in zip(grupo['codigo'], grupo['qtd_peca'], grupo['comprimento'], grupo['conjunto']):
        if codigo not in codigos_quantidades:
            codigos_quantidades[codigo] = {
                'quantidade': qtd / qtd_vara,  # Divide pela quantidade de varas
                'comprimento': comprimento,
                'conjunto': conjunto
            }
        else:
            codigos_quantidades[codigo]['quantidade'] += qtd / qtd_vara
    return codigos_quantidades

# Função para gerar os arquivos Excel com base nos resultados armazenados
def gerar_arquivos_excel(resultados_pickle_path):
    # Carregar os resultados armazenados
    with open(resultados_pickle_path, 'rb') as f:
        resultados_totais = pickle.load(f)

    # Agrupar os resultados por ordem
    df_resultados = pd.DataFrame(resultados_totais)#.to_csv("result.csv",index=False)
    df_resultados = df_resultados.sort_values('codigo')
    grupos_ordem = df_resultados.groupby('ordem_global')
    df_resultados['comprimento'] = df_resultados['tamanho_alocado_item'] / df_resultados['quantidade_alocada']
    # df_resultados[df_resultados['codigo']=='28915']
    
    # for ordem, grupo in grupos_ordem:
    #     resultados_iteracao = grupo.to_dict('records')
    #     print(resultados_iteracao)
    #     # Preencher o Excel com os resultados da ordem atual
    #     preencher_excel_ordem(excel_template_path, resultados_iteracao, ordem)

    # Adiciona a coluna de quantidade de varas ao DataFrame
    df_resultados['qt_varas'] = df_resultados['tamanho_alocado_item'] / 6000

    # Agrupa o DataFrame por 'mp' e 'ordem_global'
    grupos_mp = df_resultados.groupby(['mp', 'ordem_global'])
    
    json_l = []

    ultima_mp = None

    # Itera sobre cada grupo de matéria-prima e ordem
    for (mp, ordem), grupo in grupos_mp:
        # Calcula o total de varas para a matéria-prima se a matéria-prima atual for diferente da última
        if mp != ultima_mp:
            total_vara = grupo['qt_varas'].sum()
            total_vara_rounded = round(total_vara)
            ultima_mp = mp  # Atualiza a última matéria-prima impressa

        # Calcula a perda total de matéria-prima para a ordem atual
        perda_total = grupo['perda_materia_prima'].sum()  # Soma todas as perdas individuais da ordem

        # Itera sobre as linhas do grupo e adiciona os detalhes ao JSON
        for _, linha in grupo.iterrows():
            codigo = "0" + linha['codigo'] if len(linha['codigo']) == 5 else linha['codigo']
            json_l.append({
                'ordem': ordem,
                'codigo': codigo,
                'conjunto': linha['conjunto'],
                'qtd_peca': linha['quantidade_alocada'],
                'perda': perda_total,
                'materia_prima': mp,
                'comprimento': linha['comprimento']
            })

    # Criação do DataFrame a partir da lista JSON
    df_resultado = pd.DataFrame(json_l)

    # df_resultado['codigos_concatenados'] = df_resultado.groupby('ordem')['codigo'].transform(lambda x: ', '.join(x))

    # Criar uma nova série com códigos concatenados
    codigos_concatenados = df_resultado.groupby('ordem').apply(
        lambda x: ', '.join(f"{row['codigo']} - {row['qtd_peca']}" for _, row in x.iterrows())
    ).reset_index(name='codigos_concatenados')

    # Mesclar a nova série de volta ao DataFrame original
    df_resultado = df_resultado.merge(codigos_concatenados, on='ordem', how='left')

    df_resultado['quantidade_ordem'] = df_resultado.groupby('ordem')['ordem'].transform('count')
    df_resultado['qtd_vara'] = 1 / df_resultado['quantidade_ordem']
    df_resultado[df_resultado['ordem']==244]
    df_resultado_grouped = df_resultado.groupby('codigos_concatenados').agg({
        'ordem': 'first',  # Mantém o primeiro valor
        'materia_prima':'first',
        'codigo': list,  # Mantém todos os códigos (temporário)
        'qtd_peca': list,  # Mantém todas as quantidades (temporário)
        'conjunto': list,  # Mantém o conjunto completo
        'perda': 'sum',  # Soma as perdas
        'quantidade_ordem': 'sum',  # Soma as quantidades por ordem
        'qtd_vara': 'sum',  # Soma as varas
        'comprimento': list
    }).reset_index()

    # Aplicar consolidação dos códigos e suas quantidades
    df_resultado_grouped['codigo_quantidades'] = df_resultado_grouped.apply(
        lambda row: consolidar_codigos_quantidades(
            pd.DataFrame({
                'codigo': row['codigo'],
                'qtd_peca': row['qtd_peca'],
                'comprimento': row['comprimento'],
                'conjunto': row['conjunto']
            }),
            row['qtd_vara']
        ), axis=1
    )

    # Remover colunas temporárias (opcional)
    df_resultado_grouped = df_resultado_grouped.drop(columns=['codigo', 'qtd_peca'])

    # df_resultado[df_resultado['codigos_concatenados']=='028116, 028123']
    # Agrupa o DataFrame para somar qtd_peca e qtd_vara por matéria-prima, código e códigos concatenados
    # df_agrupado_final = df_resultado.groupby(
    #     ['materia_prima', 'conjunto', 'codigos_concatenados', 'codigo']
    # ).agg(
    #     qtd_peca_sum=('qtd_peca', 'sum'),
    #     qtd_vara_sum=('qtd_vara', 'sum')
    # )
    
    # # Exibe o DataFrame agrupado
    # df_agrupado_final = df_agrupado_final.sort_values('codigos_concatenados')

    return df_resultado_grouped

# Função para preencher a planilha Excel com os resultados da ordem
# def preencher_excel_ordem(excel_template_path, df_agrupado):
#     # Itera sobre cada grupo de códigos concatenados diferentes
#     for codigos_concatenados, grupo in df_agrupado.groupby('codigos_concatenados'):
#         # Carrega o template Excel
#         wb = load_workbook(excel_template_path)
#         ws = wb.active  # Supondo que o template tem apenas uma planilha ativa

#         start_row = 10  # Linha inicial para preenchimento
#         codigo_col = 'A'  # Coluna para os códigos
#         quantidade_col = 'C'  # Coluna para as quantidades
#         comprimento_col = 'D'  # Coluna para os comprimentos
#         qtd_planejada_col = 'E'
#         conjunto_col = 'F'  # Coluna para os conjuntos
        
#         # Extrai a matéria-prima e quantidade de vara do grupo
#         mp = str(grupo['materia_prima'].iloc[0])  # Converte para string
#         qtd_vara = str(grupo['qtd_vara'].iloc[0])  # Converte para string

#         # Define as células específicas no cabeçalho
#         ws['B4'] = f'OP-{codigos_concatenados}'
#         ws['B5'] = mp
#         ws['B6'] = '6000'
#         ws['E6'] = qtd_vara

#         # Preenche as células com os dados do grupo
#         for idx, (codigo, detalhes) in enumerate(grupo['codigo_quantidades'].iloc[0].items(), start=0):
#             row = start_row + idx
#             ws[f'{codigo_col}{row}'] = codigo  # Preenche o código
#             ws[f'{quantidade_col}{row}'] = detalhes['quantidade']  # Preenche a quantidade
#             ws[f'{comprimento_col}{row}'] = detalhes['comprimento']  # Preenche o comprimento
#             ws[f'{conjunto_col}{row}'] = detalhes['conjunto']
#             ws[f'{qtd_planejada_col}{row}'] = float(detalhes['quantidade']) * float(qtd_vara)

#         # Define o caminho de saída e salva o arquivo com um nome específico
#         output_excel_path = f"OP_{codigos_concatenados.replace(', ', '_')}.xlsx"
#         output_dir = r"C:\Users\pcp2\Downloads\teste_csv"
#         os.makedirs(output_dir, exist_ok=True)
#         wb.save(os.path.join(output_dir, output_excel_path))

#         print(f"Arquivo {output_excel_path} salvo com sucesso em {output_dir}")

# Caminho para o CSV de entrada fornecido pelo usuário
input_csv = 'input_pecas.csv'  # Substitua pelo caminho do seu arquivo CSV de entrada

# Caminho para o template Excel fornecido pelo usuário
excel_template_path = 'novo_template.xlsx'  # Substitua pelo caminho do seu template Excel

# Caminho para armazenar os resultados da otimização
resultados_pickle_path = 'resultados_otimizacao.pkl'
resultados_csv_path = 'resultado_otimizacao.csv'  # Novo arquivo CSV com as informações gerais

# Carregar os dados do CSV fornecido pelo usuário
# df_agrupado_final = carregar_dados_csv(input_csv)

# # Executar a otimização e armazenar os resultados (sem gerar os arquivos Excel)
# otimizar_e_armazenar_resultados(df, resultados_pickle_path, resultados_csv_path)

# # Em algum momento posterior, chamar a função para gerar os arquivos Excel
# gerar_arquivos_excel(resultados_pickle_path, excel_template_path)

# df_agrupado_final = gerar_arquivos_excel(resultados_pickle_path)

# preencher_excel_ordem(excel_template_path,df_agrupado_final)

def preencher_excel_ordem(excel_template_path, df_agrupado, seq):
    arquivos_gerados = []  # Lista para armazenar os arquivos gerados
    resumo_geral = []  # Lista para armazenar os dados do resumo

    # Itera sobre cada grupo de códigos concatenados diferentes
    for codigos_concatenados, grupo in df_agrupado.groupby('codigos_concatenados'):
        # Carrega o template Excel
        wb = load_workbook(excel_template_path)
        ws = wb.active  # Supondo que o template tem apenas uma planilha ativa

        start_row = 10  # Linha inicial para preenchimento
        codigo_col = 'A'  # Coluna para os códigos
        quantidade_col = 'C'  # Coluna para as quantidades
        comprimento_col = 'D'  # Coluna para os comprimentos
        qtd_planejada_col = 'E'
        conjunto_col = 'F'  # Coluna para os conjuntos

        # Extrai a matéria-prima e quantidade de vara do grupo
        mp = str(grupo['materia_prima'].iloc[0])  # Converte para string
        qtd_vara = str(grupo['qtd_vara'].iloc[0])  # Converte para string

        # Define as células específicas no cabeçalho
        ws['B4'] = f'OS-{seq}'
        ws['B5'] = mp
        ws['B6'] = '6000'
        ws['G5'] = qtd_vara
        seq += 1
        perca_por_peca = 0

        # Preenche as células com os dados do grupo
        for idx, (codigo, detalhes) in enumerate(grupo['codigo_quantidades'].iloc[0].items(), start=0):
            row = start_row + idx
            ws[f'{codigo_col}{row}'] = codigo  # Preenche o código
            ws[f'{quantidade_col}{row}'] = detalhes['quantidade']  # Preenche a quantidade
            ws[f'{comprimento_col}{row}'] = detalhes['comprimento']  # Preenche o comprimento
            ws[f'{conjunto_col}{row}'] = detalhes['conjunto']
            ws[f'{qtd_planejada_col}{row}'] = float(detalhes['quantidade']) * float(qtd_vara)
            perca_por_peca += detalhes['quantidade'] * detalhes['comprimento']

            # Adiciona os dados ao resumo geral
            resumo_geral.append({
                "OS": seq,
                "Código": codigo,
                "Quantidade": detalhes['quantidade'],
                "Comprimento": detalhes['comprimento'],
                "Conjunto": detalhes['conjunto'],
                "Qtd Planejada": float(detalhes['quantidade']) * float(qtd_vara),
                "Perca": perca_por_peca,

            })

        perca_total = 6000 - perca_por_peca
        ws['G7'] = perca_total

        # Salva o arquivo em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Armazena o arquivo em memória com o nome
        arquivos_gerados.append((f"OP_{codigos_concatenados.replace(', ', '_')}.xlsx", output))

    return arquivos_gerados, resumo_geral

st.title("Otimização de Peças e Geração de Excel")

# Upload do arquivo CSV
uploaded_file = st.file_uploader("Faça upload do arquivo CSV", type=["csv"])

st.text('Formato do arquivo: CSV, separado por virgula')
st.text('Colunas: Código,MP,Qntd,Comprimento,Conjunto')
st.text("""Exemplo:
Código,MP,Qntd,Comprimento,Conjunto
027783,110401 - ACO LAMINADO RED 2.3/4 1020,12,28,027781 - RODA DENTADA 14 DENTES - ELEV MENOR / ELEV MAIOR - SELECIONADORA DE FRUTAS
028116,134038 - TUBO QDRINOX ANSI 304 40 X 40 X 2.00,12,1750,028126 - SUP SUPERIOR - ELEV MENOR - SELECIONADORA DE FRUTAS
""")

# Processa os dados se o arquivo for carregado
if uploaded_file is not None:
    # Carregar os dados CSV como DataFrame
    df = pd.read_csv(uploaded_file, sep=',')
    
    # Mostrar os dados carregados
    st.dataframe(df)
    
seq = st.number_input('N° início da sequência:', min_value=1, value=1, step=1)

if st.button("Processar e Gerar Excel"):

    if uploaded_file is None:
        st.error("Nenhum arquivo CSV foi carregado. Por favor, faça o upload.")
    else:
        # Tratar os dados
        try:
            df_tratado = carregar_dados_csv(df)
        except:
            st.error("Erro no formato do CSV.")
    
        # Resolver otimização e salvar resultados
        otimizar_e_armazenar_resultados(df_tratado, resultados_pickle_path, resultados_csv_path)

        # Gerar arquivos Excel
        df_resultado_grouped = gerar_arquivos_excel(resultados_pickle_path)

        # Preencher os Excel e obter os arquivos gerados e o resumo
        arquivos_gerados, resumo_geral = preencher_excel_ordem(excel_template_path, df_resultado_grouped, seq)

        # Criar DataFrame do resumo para salvar como CSV
        df_resumo = pd.DataFrame(resumo_geral)
        resumo_excel_buffer = BytesIO()  # Correção no nome para clareza
        df_resumo.to_excel(resumo_excel_buffer, index=False, engine="openpyxl")  # Remove encoding e adiciona engine
        resumo_excel_buffer.seek(0)

        # Criar um arquivo ZIP em memória
        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
            for nome, excel in arquivos_gerados:
                # Adiciona cada arquivo Excel ao ZIP
                zipf.writestr(nome, excel.getvalue())
            # Adiciona o arquivo de resumo ao ZIP
            zipf.writestr("resumo_geral.xlsx", resumo_excel_buffer.getvalue())

        # Movendo o ponteiro para o início do arquivo ZIP
        zip_buffer.seek(0)

        # Disponibiliza o arquivo ZIP para download
        st.download_button(
            label="Baixar Todos os Arquivos em ZIP",
            data=zip_buffer,
            file_name="planilhas_com_resumo.zip",
            mime="application/zip",
        )
